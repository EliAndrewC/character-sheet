"""Regenerate binary happy-path fixtures from the canonical plain text.

Run this script whenever `happy_plaintext.txt` changes so the binary-format
variants stay in sync.

    python3 tests/import_fixtures/regenerate_happy_path.py

Generates: .rtf .docx .odt .xls .xlsx .ods .pdf in happy_path/, plus
.doc via LibreOffice headless if `soffice` is on PATH.

Legacy `.sxw` (pre-fork OpenOffice) cannot be written - modern LibreOffice
has dropped the export filter - and no pure-Python writer exists either.
Deferred until a real `.sxw` sample from a player is available.

See import-design/design.md §13.1 for the canonical character definition.
"""

from pathlib import Path

HERE = Path(__file__).parent
SRC = HERE / "happy_path" / "happy_plaintext.txt"
OUT = HERE / "happy_path"


def read_canonical() -> str:
    return SRC.read_text(encoding="utf-8")


# -------- RTF --------------------------------------------------------------

def write_rtf(text: str) -> None:
    # Minimal RTF: escape backslashes/braces, replace newlines with \par.
    escaped = text.replace("\\", "\\\\").replace("{", "\\{").replace("}", "\\}")
    body = escaped.replace("\n", "\\par\n")
    rtf = r"{\rtf1\ansi\deff0" + "\n" + body + "\n}"
    (OUT / "happy_rtf.rtf").write_text(rtf, encoding="utf-8")


# -------- DOCX -------------------------------------------------------------

def write_docx(text: str) -> None:
    from docx import Document
    doc = Document()
    for line in text.splitlines():
        doc.add_paragraph(line)
    doc.save(str(OUT / "happy_docx.docx"))


# -------- ODT --------------------------------------------------------------

def write_odt(text: str) -> None:
    from odf.opendocument import OpenDocumentText
    from odf.text import P
    doc = OpenDocumentText()
    for line in text.splitlines():
        doc.text.addElement(P(text=line))
    doc.save(str(OUT / "happy_odt.odt"))


# -------- XLSX -------------------------------------------------------------
# Sheet layout: each line of plaintext becomes one row, column A.
# That's crude but faithful to what an LLM would see after `openpyxl`
# flattens the sheet.  The prose "Background" / "Appearance" sections go
# into the same column.

def write_xlsx(text: str) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Character"
    for i, line in enumerate(text.splitlines(), start=1):
        ws.cell(row=i, column=1, value=line if line else None)
    wb.save(str(OUT / "happy_xlsx.xlsx"))


# -------- ODS --------------------------------------------------------------

def write_ods(text: str) -> None:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
    doc = OpenDocumentSpreadsheet()
    table = Table(name="Character")
    for line in text.splitlines():
        row = TableRow()
        cell = TableCell()
        cell.addElement(P(text=line))
        row.addElement(cell)
        table.addElement(row)
    doc.spreadsheet.addElement(table)
    doc.save(str(OUT / "happy_ods.ods"))


# -------- XLS (legacy) ----------------------------------------------------

def write_xls(text: str) -> None:
    import xlwt
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Character")
    for i, line in enumerate(text.splitlines()):
        ws.write(i, 0, line)
    wb.save(str(OUT / "happy_legacy_xls.xls"))


# -------- DOC (legacy, via LibreOffice) -----------------------------------

def write_legacy_doc() -> None:
    """Produce happy_legacy_doc.doc by converting happy_docx.docx through
    soffice. Silently skips (with a warning) if soffice is not installed.
    """
    import shutil
    import subprocess

    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        print("  skip: soffice/libreoffice not on PATH - .doc not regenerated")
        return

    docx = OUT / "happy_docx.docx"
    subprocess.run(
        [soffice, "--headless", "--convert-to", "doc",
         str(docx), "--outdir", str(OUT)],
        check=True,
        capture_output=True,
    )
    produced = OUT / "happy_docx.doc"
    target = OUT / "happy_legacy_doc.doc"
    if target.exists():
        target.unlink()
    produced.rename(target)


# -------- PDF --------------------------------------------------------------

def write_pdf(text: str) -> None:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    c = canvas.Canvas(str(OUT / "happy_pdf_text.pdf"), pagesize=letter)
    _, height = letter
    y = height - 50
    for line in text.splitlines():
        if y < 50:
            c.showPage()
            y = height - 50
        c.drawString(50, y, line[:100])
        y -= 14
    c.save()


def main() -> None:
    text = read_canonical()
    write_rtf(text)
    write_docx(text)
    write_odt(text)
    write_xlsx(text)
    write_xls(text)
    write_ods(text)
    write_pdf(text)
    write_legacy_doc()  # depends on happy_docx.docx having just been written
    print("Regenerated:")
    for p in sorted(OUT.iterdir()):
        if p.suffix in {".rtf", ".doc", ".docx", ".odt", ".xlsx", ".xls", ".ods", ".pdf"}:
            print(f"  {p.relative_to(HERE.parent)} ({p.stat().st_size} bytes)")


if __name__ == "__main__":  # pragma: no cover
    main()
